# -*- coding: utf-8 -*-
"""
Created on Tue Feb  9 09:56:15 2021

@author: 皓
"""
import os.path as osp
import time
from openpyxl import Workbook, load_workbook
# from rotating_platform import RotatingPlatform
from pychariot import Chariot, RetureCode
import configuration as config


WAIT_TIME = config.max_wait_time


def create_test(api):
    global WAIT_TIME
    test_handle = api.test_new()
    pairs = []
    for conf in config.pairs_conf:
        pair = api.create_pair_attr(*conf)
        api.test_add_pair(test_handle, pair)
        pairs.append(pair)
    runopts_conf = getattr(config, 'runopts_conf', None)
    if runopts_conf:
        runopts = api.test_get_runopts(test_handle)
        if 'test_end' in runopts_conf:
            api.runopts_set_test_end(runopts, runopts_conf.get('test_end'))
        if 'test_duration' in runopts_conf:
            test_duration = runopts_conf.get('test_duration')
            WAIT_TIME = config.max_wait_time + test_duration
            api.runopts_set_test_duration(runopts, test_duration)
    return test_handle


def copy_swap_pairs_test(api, src_test):
    test_handle = api.get_swap_pairs_test(src_test)
    runopts_conf = getattr(config, 'runopts_conf', None)
    if runopts_conf:
        runopts = api.test_get_runopts(test_handle)
        if 'test_end' in runopts_conf:
            api.runopts_set_test_end(runopts, runopts_conf.get('test_end'))
        if 'test_duration' in runopts_conf:
            test_duration = runopts_conf.get('test_duration')
            api.runopts_set_test_duration(runopts, test_duration)
    return test_handle


def run_test_result(api, test_handle, wait_time):
    api.test_start(test_handle)
    ret = api.wait_for_test(test_handle, wait_time, config.timeout)
    if ret is False:
        raise TimeoutError('测试等待超时.')
    pairs = api.get_pairs(test_handle)
    print('测试发送bytes：{}，接收bytes：{}'.format(
        api.get_pairs_bytes_sent_e1(pairs),
        api.get_pairs_bytes_recv_e1(pairs)
        ))
    result = api.get_pairs_results_average(pairs)
    return result


def rotation(device, angle, retry=10):
    for i in range(retry):
        res = device.jop_to_angle(angle)
        if abs(res - angle) < 0.2:
            return
    if abs(res - angle) > 0.5:
        raise ValueError(
            '转台误差过大，设置：{}，实际：{}'.format(angle, res))


def check_ws(ws):
    if ws.max_row == 1 and ws.max_column == 1:
        ws['A1'] = ws.title
    if not ws['A1'].value == ws.title:
        raise ValueError('Result file format error.')


def set_distinct_sort(src):
    """
    :param list1: 列表
    :return: 返回列表去重后保持原来元素的顺序不变
    """
    dst = list(set(src))
    dst.sort(key=src.index)
    return dst


def init_row(ws):
    if ws.max_row > 1:
        cols = ws['A2':'A{}'.format(ws.max_row)]
        cols_value = [x[0].value for x in cols]
    else:
        cols_value = []
    cols_value.extend(config.rotation_angle)
    sort_list = set_distinct_sort(cols_value)
    res = {}
    for i, item in enumerate(sort_list):
        row = 2 + i
        ws['A{}'.format(row)] = item
        res[item] = row
    return res


def init_xlsx():
    result_info = {}
    result_path = osp.realpath(config.result_file)
    _ = result_info.setdefault('path', result_path)
    if osp.exists(result_path):
        wb = load_workbook(result_path)
    else:
        wb = Workbook()
    if 'TX' not in wb:
        tx_ws = wb.create_sheet()
        tx_ws.title = "TX"
        tx_ws['A1'] = 'TX'
    else:
        tx_ws = wb['TX']
    if 'RX' not in wb:
        rx_ws = wb.create_sheet()
        rx_ws.title = "RX"
        rx_ws['A1'] = 'RX'
    else:
        rx_ws = wb['RX']
    check_ws(tx_ws)
    check_ws(rx_ws)
    _ = result_info.setdefault('tx_rows', init_row(tx_ws))
    _ = result_info.setdefault('tx_col', tx_ws.max_column + 1)
    _ = result_info.setdefault('rx_rows', init_row(rx_ws))
    _ = result_info.setdefault('rx_col', rx_ws.max_column + 1)
    _ = result_info.setdefault('tx_ws', tx_ws)
    _ = result_info.setdefault('rx_ws', rx_ws)
    _ = result_info.setdefault('workbook', wb)
    wb.save(result_path)
    return result_info


def main():
    init_xlsx()
    api = Chariot()
    ret = api.api_initialize()
    if not ret == RetureCode.CHR_OK:
        raise RuntimeError('Chariot api initialize error.')
    ret = api.api_get_version()
    print('chrapi version:{}'.format(ret))
    tx_test = create_test(api)
    rx_test = copy_swap_pairs_test(api, tx_test)
    result_info = init_xlsx()
    # device = RotatingPlatform(config.rotating_platform)
    wb = result_info.get('workbook')
    result_path = result_info.get('path')
    tx_rows = result_info.get('tx_rows')
    rx_rows = result_info.get('rx_rows')
    tx_col = result_info.get('tx_col')
    rx_col = result_info.get('rx_col')
    tx_ws = result_info.get('tx_ws')
    rx_ws = result_info.get('rx_ws')
    for angle in config.rotation_angle:
        print('-测试转台角度：{}'.format(angle))
        # rotation(device, angle)
        print('--测试TX')
        ret = run_test_result(api, tx_test, WAIT_TIME)
        tx_ws.cell(row=tx_rows.get(angle), column=tx_col).value = round(ret, 3)
        print('--测试RX')
        ret = run_test_result(api, rx_test, WAIT_TIME)
        rx_ws.cell(row=rx_rows.get(angle), column=rx_col).value = round(ret, 3)
        wb.save(result_path)
    else:
        time_text = time.strftime("%Y/%m/%d %H:%M:%S", time.localtime())
        tx_ws.cell(row=1, column=tx_col).value = time_text
        rx_ws.cell(row=1, column=rx_col).value = time_text
        wb.save(result_path)


if __name__ == '__main__':
    main()
